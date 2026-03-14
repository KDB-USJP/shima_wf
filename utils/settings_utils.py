import json
import os
from pathlib import Path

class ShimaSettings:
    """Utility to load and cache centralized site settings with fallbacks."""
    
    _config = None
    _config_path = Path(__file__).parent.parent / "config" / "site_default_settings.json"
    _user_config = None
    _user_config_path = Path(__file__).parent.parent / "config" / "shima_settings.json"
    _excel_palettes = None

    # --- FALLBACKS ---
    DEFAULT_API_BASE = "https://shima.wf"
    DEFAULT_PACKS = {
        "Walking Woman": "https://drive.google.com/uc?export=download&id=1nHhDNMWj3Reox5yfo-zlk_4Ry4KsC7hZ",
        "Still Life": "https://drive.google.com/uc?export=download&id=1oEQi6lnq32D9thOUEySL1NtYk66pDmpt"
    }

    DEFAULT_COMMONS = {
        "model_types": [
            "sdxl", "sd1.5", "sd2.x", "sd3",
            "flux", "pony", "illustrious",
            "auraflow", "hunyuan",
            "lumina2", "chroma", "hidream",
            "z-image-base", "z-image-turbo"
        ],
        "aspect_ratios": [
            "1:1 Square", 
            "16:9 Widescreen", 
            "4:3 Standard", 
            "21:9 Ultrawide", 
            "3:2 Photo",
            "Custom"
        ],
        "orientations": ["landscape", "portrait", "auto"]
    }

    DEFAULT_MULTISAVER = {
        "filename_order_presets": [
            "PRE,PRJ,BN,ET,SUF,TS,CID",
            "BN,ET,TS,CID"
        ],
        "separators": ["_", "-", ".", " ", ""]
    }

    @classmethod
    def get_excel_palettes(cls):
        """Parse E:\ComfyDev\Shima\assets\data\shima_sheets.xlsx (tab: node-color-themes) once."""
        if cls._excel_palettes is not None:
            return cls._excel_palettes
            
        # Dynamically resolve path relative to this file
        # __file__ is in utils/, so go up one level to Shima root, then into assets/data
        excel_path = Path(__file__).parent.parent / "assets" / "data" / "shima_sheets.xlsx"
        palettes = {}
        
        if not excel_path.exists():
            cls._excel_palettes = palettes
            return palettes

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            if "node-color-themes" not in wb.sheetnames:
                return palettes
            
            sheet = wb["node-color-themes"]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return palettes

            # Header row (starting col 2) has theme names
            header = rows[0]
            theme_names = [name for name in header[1:] if name]
            
            # Initialize palettes
            for name in theme_names:
                palettes[name] = {"node": {}}

            # Data rows
            for row in rows[1:]:
                node_key = row[0]
                if not node_key:
                    continue
                
                for idx, color in enumerate(row[1:]):
                    if idx < len(theme_names) and color:
                        theme_name = theme_names[idx]
                        palettes[theme_name]["node"][node_key] = color

            cls._excel_palettes = palettes
            return palettes
        except Exception as e:
            print(f"[Shima] Error parsing Excel palettes: {e}")
            cls._excel_palettes = palettes
            return palettes

    @classmethod
    def get_config(cls):
        if cls._config is None:
            cls.reload()
        
        # Merge Excel palettes into config
        config = cls._config.copy()
        if "themes" not in config:
            config["themes"] = {}
        
        excel_palettes = cls.get_excel_palettes()
        if excel_palettes:
            if "palettes" not in config["themes"]:
                config["themes"]["palettes"] = {}
            config["themes"]["palettes"].update(excel_palettes)
            
        return config

    @classmethod
    def get_user_config(cls):
        # ... rest of the file ...
        if cls._user_config is None:
            cls.reload_user()
        return cls._user_config

    @classmethod
    def reload(cls):
        if cls._config_path.exists():
            try:
                with open(cls._config_path, "r") as f:
                    cls._config = json.load(f)
                    print(f"[Shima] Loaded site settings from {cls._config_path}")
            except Exception as e:
                print(f"[Shima] Error loading site settings: {e}")
                cls._config = {}
        else:
            cls._config = {}

    @classmethod
    def reload_user(cls):
        if cls._user_config_path.exists():
            try:
                with open(cls._user_config_path, "r") as f:
                    cls._user_config = json.load(f)
                    print(f"[Shima] Loaded user settings from {cls._user_config_path}")
            except Exception as e:
                print(f"[Shima] Error loading user settings: {e}")
                cls._user_config = {}
        else:
            cls._user_config = {}

    @classmethod
    def get_api_base(cls):
        """Returns the server URL from user settings, falling back to default."""
        return cls.get_user_config().get("api_base", cls.DEFAULT_API_BASE)

    @classmethod
    def get_asset_packs(cls):
        return cls.get_config().get("asset_packs", cls.DEFAULT_PACKS)

    @classmethod
    def get_commons(cls):
        return cls.get_config().get("commons", cls.DEFAULT_COMMONS)

    @classmethod
    def get_multisaver(cls):
        return cls.get_config().get("multisaver", cls.DEFAULT_MULTISAVER)

    @classmethod
    def get_list(cls, section, key, default):
        """Helper to safely get a list from a section."""
        return cls.get_config().get(section, {}).get(key, default)
