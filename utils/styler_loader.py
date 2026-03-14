import os
import csv
import json
import openpyxl

class StylerDataLoader:
    """
    Loads style data from a multi-sheet XLSX file.
    Sheets: "artists" (Index, Name, OptionalNegative, Categories, Info)
    Sheets: "your_styles" (Positive, Negative, Categories, Info)
    """
    def __init__(self, file_path, assets_dir=None):
        self.file_path = file_path
        self.assets_dir = assets_dir
        self.artists_data = []
        self.user_data = [] # New list for "your_styles"
        self.categories = set()
        self.load_data()

    def load_data(self):
        # Fallback to TSV if XLSX missing? Or assume user provided it? 
        # User said: "I have remade the data... shima_sheets.xlsx"
        # We should only support XLSX now based on this refactor.
        
        if not os.path.exists(self.file_path):
            # print(f"[ShimaStyler] Warning: Data file not found at {self.file_path}")
            return

        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            
            # --- Load Artists ---
            if "artists" in wb.sheetnames:
                sheet = wb["artists"]
                # iter_rows yields tuples. min_row=2 skips header.
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Columns: FirstName(0), LastName(1), optional_negative(2), Categories(3), Extrainfo(4)
                    if not row or (row[0] is None and row[1] is None): continue
                    
                    first = str(row[0]).strip() if row[0] is not None else ""
                    last = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                    
                    # Combine names handling empty/None
                    if first and last:
                        name = f"{first} {last}"
                    else:
                        name = first or last
                        
                    if not name: continue

                    opt_neg = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                    cats_str = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                    info = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                    
                    # Store categories
                    row_cats = [c.strip() for c in cats_str.split(',') if c.strip()]
                    for c in row_cats: self.categories.add(c)
                    
                    self.artists_data.append({
                        "id": f"A{len(self.artists_data)}",
                        "name": name,
                        "type": "artist",
                        "positive": name, # Base positive prompt is just the name
                        "negative": opt_neg,
                        "categories": row_cats,
                        "info": info
                    })
                # print(f"[ShimaStyler] Loaded {len(self.artists_data)} artists.")
            else:
                pass
                # print("[ShimaStyler] 'artists' sheet not found in XLSX.")

            # --- Load User Styles ---
            if "your_styles" in wb.sheetnames:
                sheet = wb["your_styles"]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        # Columns: style_name(0), positive(1), negative(2), Categories(3), Extrainfo(4)
                        if not row: continue
                        
                        # Safe access helper
                        def get_col(idx):
                            if idx < len(row) and row[idx] is not None:
                                return str(row[idx]).strip()
                            return ""

                        style_name = get_col(0)
                        pos = get_col(1)
                        
                        if not style_name and not pos: continue
                        
                        # If name is missing but pos exists, use truncated pos as label (fallback)
                        if not style_name:
                            style_name = pos
                            if len(style_name) > 40:
                                style_name = style_name[:37] + "..."

                        neg = get_col(2)
                        cats_str = get_col(3)
                        info = get_col(4)
                        
                        row_cats = [c.strip() for c in cats_str.split(',') if c.strip()]
                        for c in row_cats: self.categories.add(c)
                        
                        self.user_data.append({
                            "id": f"U{len(self.user_data)}",
                            "name": style_name,
                            "type": "user_style",
                            "positive": pos,
                            "negative": neg,
                            "categories": row_cats,
                            "info": info
                        })
                    except Exception as row_err:
                        # print(f"[ShimaStyler] Error parsing user style row: {row_err}")
                        continue
                        
                # print(f"[ShimaStyler] Loaded {len(self.user_data)} user styles.")

        except Exception as e:
            pass
            # print(f"[ShimaStyler] Error loading XLSX: {e}")

    def get_data(self):
        # Return structured data for frontend tabs
        return {
            "artists": self.artists_data,
            "user_styles": self.user_data
        }

    def get_categories(self):
        return sorted(list(self.categories))
